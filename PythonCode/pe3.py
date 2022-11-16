from openpyxl import Workbook
wb = Workbook()
#GRAB THE ACTIVE WORKSHEET
ws = wb.active
#active sheet is that is open or specify the name of workbook
#ASSIGNING DATA DIRECTLY TO CELLS
ws['A1'] = 'Invoice.no'
ws['B1'] = 'stock code'
ws['C1'] = 'Description'
ws['D1'] = 'Quantity'
ws['E1'] = 'InvoiceData'
ws['F1'] = 'UnitPrice'
ws['G1'] = 'CustomerId'
ws['H1'] = 'Country'
#ROWS CAN ALSO BE APPENDED
ws.append([536162,'65463A8','WHITE HANGING HEART',6,'01-12-2010 087:06',2.55,17850,'United Kingdom'])
ws.append([536163,'65763A8','WHITE HANGING HEART',7,'01-12-2010 087:06',4.85,17851,'United Kingdom'])
ws.append([536164,'65464A8','WHITE HANGING HEART',2,'01-12-2010 087:06',7.55,17852,'United states'])
ws.append([536165,'62563A8','WHITE HANGING HEART',8,'01-12-2010 087:06',8.35,17853,'United states'])

wb.save("online_retail.csv")