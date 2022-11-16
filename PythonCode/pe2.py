from openpyxl import Workbook
wb = Workbook()
#GRAB THE ACTIVE WORKSHEET
ws = wb.active
#active sheet is that is open or specify the name of workbook
#ASSIGNING DATA DIRECTLY TO CELLS
ws['A1'] = 42
#ROWS CAN ALSO BE APPENDED
ws.append([1,2,3])

import datetime

ws['A2'] = datetime.datetime.now()

wb.save("sample.xlsx")